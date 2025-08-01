import streamlit as st
import streamlit_authenticator as stauth
import pickle
from pathlib import Path
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings, HuggingFaceInstructEmbeddings
from langchain.vectorstores import FAISS
from langchain.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from htmlTemplates import css, bot_template, user_template
# import mysql.connector
import openpyxl

st.set_page_config(page_title="AcademiaChat", page_icon="📚")


# def get_pdf_text(pdf_docs):                                                                                                                                                                                                    
#     text = ""
#     for pdf in pdf_docs:
#         pdf_reader = PdfReader(pdf)
#         for page in pdf_reader.pages:
#             text += page.extract_text()
#     return text


def get_excel_text(file_path):
    text = ""
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate over all rows in the sheet
        for row in sheet.iter_rows(values_only=True):
            # Concatenate text from each cell in the row
            row_text = " ".join(str(cell) for cell in row if cell is not None)
            text += row_text + "\n"  # Add a newline between rows

    return text


def get_text_chunks(raw_text):
    text_splitter = CharacterTextSplitter(
        separator="\n", chunk_size=1000, chunk_overlap=200, length_function=len
    )
    chunks = text_splitter.split_text(raw_text)
    return chunks


def get_vectorstore(text_chunks):
    embeddings = OpenAIEmbeddings()
    vectorstore = FAISS.from_texts(texts=text_chunks, embedding=embeddings)
    return vectorstore

                
def get_conversation_chain(vectorstore):
    llm = ChatOpenAI()

    memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=llm, retriever=vectorstore.as_retriever(), memory=memory
    )
    return conversation_chain


def handle_userinput(user_question):
    response = st.session_state.conversation({"question": user_question})

    st.session_state.chat_history = response["chat_history"]
    for i, message in enumerate(st.session_state.chat_history):
        if i % 2 == 0:
            st.write(
                user_template.replace("{{MSG}}", message.content),
                unsafe_allow_html=True,
            )
        else:
            st.write(
                bot_template.replace("{{MSG}}", message.content),
                unsafe_allow_html=True,
            )


names = ["Aman Dev", "Murli Dhar"]
usernames = ["vtu19464", "tts1234"]
# load hashed passwords
file_path = Path(__file__).parent / "hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_passwords = pickle.load(file)

authenticator = stauth.Authenticate(
    names,
    usernames,
    hashed_passwords,
    "chatbot",
    "abcdef",
    cookie_expiry_days=0,
)
name, authentication_status, username = authenticator.login("Login", "main")
identity = username[0:3]


# mydb = mysql.connector.connect(
#     host="localhost", user="root", password="Amanbondla2003.", database="chatbot"
# )
# mycursor = mydb.cursor()
# mycursor.execute("select * from student")
# result = mycursor.fetchall()
# for row in result:
#     data = row

if authentication_status:
    st.success(f"Welcome *{name}*")

    def main():
        load_dotenv()
        st.write(css, unsafe_allow_html=True)

        if "conversation" not in st.session_state:
            st.session_state.conversation = None
        if "chat_history" not in st.session_state:
            st.session_state.chat_history = None

        st.header("AcademiaChat📚")
        user_question = st.text_input("Ask me anything about your acadamics:")
        if user_question:
            handle_userinput(user_question)
        if identity == "vtu":
            file_path = (
                "C:\\Users\\USER\\Desktop\\Work\\Projects\\AcademiaChat\\student.xlsx"
            )

        elif identity == "tts":
            file_path = ("C:\\Users\\Desktop\\Work\\Projects\\AcademiaChat\\teacher.xlsx")

        raw_text = get_excel_text(file_path)
        text_chunks = get_text_chunks(raw_text)
        vectorstore = get_vectorstore(text_chunks)
        st.session_state.conversation = get_conversation_chain(vectorstore)
        authenticator.logout("Logout", "main")
    if __name__ == "__main__":
        main()
    
if authentication_status == False:
    st.error("Username/password is incorrect")
if authentication_status == None:
    st.warning("Please enter your username and password")
